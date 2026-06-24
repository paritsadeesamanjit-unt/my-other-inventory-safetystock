"""
export_utils.py — สร้าง Excel รายงานสรุป
"""
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter


HEADER_BLUE = "1A5276"
HEADER_TEAL = "1A7A6E"
HEADER_ORANGE = "935116"
HEADER_DARK = "1C2833"
ROW_ALT = "F2F3F4"
WHITE = "FFFFFF"


def _apply_header(ws, row_num: int, color_hex: str):
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[row_num]:
        if cell.value is not None:
            cell.font = Font(name="Arial", bold=True, size=10, color=WHITE)
            cell.fill = PatternFill("solid", start_color=color_hex)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border


def _apply_rows(ws, start_row: int, end_row: int):
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for i, row in enumerate(ws.iter_rows(min_row=start_row, max_row=end_row)):
        fill_color = ROW_ALT if i % 2 == 0 else WHITE
        for cell in row:
            cell.fill = PatternFill("solid", start_color=fill_color)
            cell.font = Font(name="Arial", size=9)
            cell.border = border
            cell.alignment = Alignment(vertical="center")


def _set_col_widths(ws, widths: list):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def export_full_report(
    true_stock: pd.DataFrame,
    pr_po_adj: pd.DataFrame,
    usage_summary: pd.DataFrame,
    expiry_alerts: pd.DataFrame,
    stock_alerts: pd.DataFrame,
) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: Stock Summary ────────────────────────────────────────────────
    ws1 = wb.create_sheet("Stock Summary")
    ws1.freeze_panes = "A2"
    if not true_stock.empty:
        cols = ["Material_ID", "Material_Name", "Unit", "Sources",
                "SAP_Unrestricted", "SAP_Inspection", "SAP_InTransit",
                "NoSAP_Remaining", "Total_Stock"]
        df = true_stock[cols].copy()
        ws1.append(cols)
        for _, r in df.iterrows():
            ws1.append(r.tolist())
        _apply_header(ws1, 1, HEADER_BLUE)
        _apply_rows(ws1, 2, ws1.max_row)
        _set_col_widths(ws1, [20, 50, 8, 20, 16, 14, 12, 16, 14])
        ws1.auto_filter.ref = ws1.dimensions

    # ── Sheet 2: PR/PO Adjusted ───────────────────────────────────────────────
    ws2 = wb.create_sheet("PR-PO Adjusted")
    ws2.freeze_panes = "A2"
    if not pr_po_adj.empty:
        keep = ["PR_Number", "PO_Number", "Material_ID", "Product_Spec",
                "Unit", "Vendor", "PR_Date", "Require_Date",
                "Required_QTY", "Accu_Rece_QTY", "Outstanding_QTY",
                "Adjusted_Outstanding"]
        keep = [c for c in keep if c in pr_po_adj.columns]
        df2 = pr_po_adj[keep].copy()
        ws2.append(keep)
        for _, r in df2.iterrows():
            ws2.append([
                str(v) if pd.isna(v) == False and not isinstance(v, (int, float)) else v
                for v in r.tolist()
            ])
        _apply_header(ws2, 1, HEADER_TEAL)
        _apply_rows(ws2, 2, ws2.max_row)
        _set_col_widths(ws2, [14, 14, 20, 45, 8, 35, 14, 14, 14, 14, 14, 18])
        ws2.auto_filter.ref = ws2.dimensions

    # ── Sheet 3: Usage Summary ────────────────────────────────────────────────
    ws3 = wb.create_sheet("Usage Summary")
    ws3.freeze_panes = "A2"
    if not usage_summary.empty:
        ws3.append(usage_summary.columns.tolist())
        for _, r in usage_summary.iterrows():
            ws3.append(r.tolist())
        _apply_header(ws3, 1, HEADER_DARK)
        _apply_rows(ws3, 2, ws3.max_row)
        _set_col_widths(ws3, [20, 50, 8, 14, 16, 14])
        ws3.auto_filter.ref = ws3.dimensions

    # ── Sheet 4: Expiry Alerts ────────────────────────────────────────────────
    ws4 = wb.create_sheet("Expiry Alerts")
    ws4.freeze_panes = "A2"
    if not expiry_alerts.empty:
        ws4.append(expiry_alerts.columns.tolist())
        for _, r in expiry_alerts.iterrows():
            ws4.append(r.tolist())
        _apply_header(ws4, 1, HEADER_ORANGE)
        _apply_rows(ws4, 2, ws4.max_row)
        _set_col_widths(ws4, [20, 50, 8, 14, 16, 16, 18, 22])
        ws4.auto_filter.ref = ws4.dimensions

    # ── Sheet 5: Stock Alerts ─────────────────────────────────────────────────
    ws5 = wb.create_sheet("Stock Alerts")
    ws5.freeze_panes = "A2"
    if not stock_alerts.empty:
        ws5.append(stock_alerts.columns.tolist())
        for _, r in stock_alerts.iterrows():
            ws5.append(r.tolist())
        _apply_header(ws5, 1, "7B241C")
        _apply_rows(ws5, 2, ws5.max_row)
        _set_col_widths(ws5, [20, 45, 14, 14, 14, 14, 14, 14])
        ws5.auto_filter.ref = ws5.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
